import os
import openpyxl
from openpyxl.styles import Color, Alignment, PatternFill, Font, Border, Side

#Changes directory
os.chdir(os.getcwd())

#loads an excel sheet
wb  = openpyxl.load_workbook("rl_trades.xlsx")

#Chooses a sheet
sellSheet = wb['Sell']
buySheet = wb['Buy']
mixSheet = wb['All']

#Styles
blueFill = PatternFill(start_color='89D8F5',
                   end_color='89D8F5',
                   fill_type='solid')

redFill = PatternFill(start_color='FF8295',
                   end_color='FF8295',
                   fill_type='solid')

solidRedFill = PatternFill(start_color='e02f2f',
                   end_color='e02f2f',
                   fill_type='solid')

solidGreenFill = PatternFill(start_color='2fe084',
                   end_color='2fe084',
                   fill_type='solid')

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


def save():
    os.chdir("D:\\Editing\\Programming\\Python Projects\\Rocket League Excel Sheet")
    wb.save('RocketLeagueTrades_BACKUP.xlsx')
    os.chdir("C:\\Users\\Tomer\\Desktop")
    wb.save('RocketLeagueTrades.xlsx')

def addTrade():
    #Input
    traderName = input("Name: ")
    BuyOrSell = input("Buy or sell: ")
    offer1 = input("What item/offer: ")
    offer2 = input("For: ")
    friend = input("Friends?: ")



    #Setup
    currentRow = mixSheet.max_row + 1
    maxColumn = mixSheet.max_column + 1 
    BuyOrSell = BuyOrSell.upper()

    #Style
    if BuyOrSell == "S":
        for i in range(1, maxColumn):
            mixSheet.cell(row=currentRow, column=i).fill = redFill
            mixSheet.cell(row=currentRow, column=i).border = thin_border
    elif BuyOrSell == "B":
        for i in range(1, maxColumn):
            mixSheet.cell(row=currentRow, column=i).fill = blueFill
            mixSheet.cell(row=currentRow, column=i).border = thin_border

        
    #Trader Name
    currentCell = mixSheet.cell(row=currentRow, column=1)
    currentCell.value = traderName
    currentCell.alignment = Alignment(horizontal='center')


    #Sell Or Buy
    currentCell = mixSheet.cell(row=currentRow, column=2)
    currentCell.value = "[" + BuyOrSell + "]"
    currentCell.alignment = Alignment(horizontal='center')

    #Offer 1
    currentCell = mixSheet.cell(row=currentRow, column=3)
    currentCell.value = offer1
    currentCell.alignment = Alignment(horizontal='center')

    #For column
    currentCell = mixSheet.cell(row=currentRow, column=4)
    currentCell.value = "FOR"
    currentCell.alignment = Alignment(horizontal='center')


    #Offer 2
    currentCell = mixSheet.cell(row=currentRow, column=5)
    if  offer2.isdigit(): 
       currentCell.value = offer2 + "c"
    else:
        currentCell.value = offer2
    
    currentCell.alignment = Alignment(horizontal='center')


    #Friends with
    currentCell = mixSheet.cell(row=currentRow, column=6)
    if friend == "n":
        currentCell.value = "NO"
        currentCell.fill = solidRedFill
        currentCell.alignment = Alignment(horizontal='center')
    elif friend == "y":
        currentCell.value = "YES"
        currentCell.fill = solidGreenFill
        currentCell.alignment = Alignment(horizontal='center')



    #Commits the changes (twice)
    try:
        save()
    except:
        inp = input("Couldn't save the changes, Try closing the Excel file. ")
        if inp == "":
            save()

#Inputs
firstInput = input("Type Enter to Start... ")
while firstInput == "":
    addTrade()
    firstInput = input("\nEnter to keep going...")
